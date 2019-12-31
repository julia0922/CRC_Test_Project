

import openpyxl
import os
from common.operation_file import OperationFile

reslultdir="E:/20人录音第一批交付-全品类/"
excelName="讯飞测试结果-全品类.xlsx"

of_obj=OperationFile()

workbook = openpyxl.Workbook()
worksheet = workbook.create_sheet(title="识别率")
worksheet1 = workbook.create_sheet(title="识别汇总")
worksheet.append(["序号", "正确文本", "ASR识别结果", "识别率"])
worksheet1.append(["语料","总数","成功","失败","成功率"])
childdirlist=os.listdir(reslultdir)
for filename in childdirlist:
    if filename.endswith(".txt")==False:
        continue

    filpath=reslultdir+"/"+filename
    alllines = of_obj.read_fileinfo(filpath)
    total=0
    success=0
    i=0
    for one_line in alllines:
        print(one_line)
        if i == 0 or one_line.find("recognition result") != -1:
            i = i + 1
            continue

        total=total+1
        asrvalue=""
        line_list = one_line.split(" ")
        if len(line_list)>4:
            for r in range(2,len(line_list)-1):
                asrvalue=asrvalue+line_list[r]
        else:
            asrvalue=line_list[2]

        result_value=str(line_list[-1])
        if result_value.lower().find("true")!=-1:
            success=success+1

        print("成功:",success)
        worksheet.append([line_list[0],line_list[1],asrvalue,result_value])
    sucess_rate = (success / total) * 100
    sucess_rate_value = "{:.2f}%".format(sucess_rate)
    worksheet1.append([filename,total,success,total-success,sucess_rate_value])

workbook.save(reslultdir+excelName)

