#!/usr/bin/env python
# coding=utf-8
### websocket模块https://pypi.python.org/pypi/websocket-client
### wavfile是一个speex压缩的ogg文件,文件格式是Ogg data, Speex audio
import websocket
from gevent import os
from websocket import ABNF
import time
import _thread
import base64
import hashlib
import os
import json
import openpyxl
from multiprocessing import Process,Manager
from xunfei.lockfile import FileLock

BASE_URL = "wss://wsapi.xfyun.cn/v1/aiui";
APPID = "5e017c12";
APIKEY = "b1c9db5797d7bd16567bb2e6a34cf075"
END_FLAG = "--end--"
param = "{\"result_level\":\"plain\",\"auth_id\":\"894c985bf8b1111c6728db79d3479aef\",\"data_type\":\"audio\",\"aue\":\"raw\",\"scene\":\"main_box\",\"sample_rate\":\"16000\",\"interact_mode\":\"oneshot\"}";

RESULT_ROOT_PATH="D:/audio_file/测试结果/讯飞/全品类/"    # 测试结果保存的路径
AUDIO_ROOT_PATH="D:/audio_file/wav音频/全品类语料/"   #音频文件存储的根路径
MANUAL_CASE_PATH="D:/audio_file/测试案例/讯飞/全品类/"  #讯飞用例文件所在路径

'''
RUN_LIST={"全品类":{"20人录音第一批交付-全品类": ["/20人录音第一批交付-全品类/","20人录音第一批交付-全品类.xlsx","true"],
               "20人录音第二批交付-全品类": ["/20人录音第二批交付-全品类/","20人录音第二批交付-全品类.xlsx","true"],
               "美的第一批递交-全品类": ["/美的第一批递交/全品类/","美的第一批递交-全品类.xlsx","true"],
               "美的第二批10人全品类": ["/美的第二批10人全品类/","美的第二批10人全品类.xlsx","true"],
               "美的第三批全品类20人": ["/美的第三批全品类20人/""true","美的第三批全品类20人.xlsx","true"],
               "美的第四批全品类10人": ["/美的第四批/","美的第四批全品类10人.xlsx","true"],
               "第五批-全品类20人": ["/第五批-全品类20人/","第五批-全品类20人.xlsx","true"]}}
'''

RUN_LIST={"全品类":{"20人录音第一批交付-全品类": ["/20人录音第一批交付-全品类/","20人录音第一批交付-全品类.xlsx","true"]}}

GEN_METHOD="全品类"
THREAD_NUM=20

def write_fileinfo(filepath, lines):
    fileobj = FileLock(filepath)
    fileobj.acquire()
    with open(filepath, 'a+', encoding="utf-8") as f:
        f.writelines(lines)
        f.writelines("\n")
    fileobj.release()

def read_fileinfo(filepath):
    list=[]
    if os.path.exists(filepath)==True:
        with open(filepath, 'r', encoding="utf-8") as f:
            for line in f:
                firstvalue = line.split(" ")[0]
                print(firstvalue)
                if firstvalue != "wav":
                    list.append(firstvalue)
    return list

def save(row):
    all_lines = " ".join(str(i) for i in row)
    if os.path.exists(save_resultpath):
        write_fileinfo(save_resultpath,lines=all_lines)
    else:
        wavinfo= " ".join(str(i) for i in ["wav", "expect", "asr", "recognition result"])
        write_fileinfo(save_resultpath, lines=wavinfo)
        write_fileinfo(save_resultpath, lines=all_lines)

    '''
    if os.path.exists(resultpath):
        wb=openpyxl.load_workbook(FILE_PATH)
        if sheetname not in wb.get_sheet_names():
            ws=wb.create_sheet(sheetname)
            ws.append(["Recognition success rate"])
            ws.append(["wav","expect","asr","recognition result"])
        else:
            ws=wb[sheetname]
    else:
        wb=openpyxl.Workbook()
        ws=wb.create_sheet(sheetname)
        ws.append(["Recognition success rate"])
        ws.append(["wav", "expect", "asr", "recognition result"])
    ws.append(row)
    wb.save(FILE_PATH)
    '''

def on_message(ws, message):
   print ('Message>>>>>>>>>>>>>>>>>>>>>>>>')
   print(message)
   if 'iat' in message:
       mJson = json.loads(message)
       resultid = mJson['data']['result_id']
       #if resultid == 1:
       result_txt= mJson['data']['text']
       basename=os.path.basename(audio_file_path)
       expectvalue=""
       result_value="FALSE"
       if basename in expectresultinfo:
           expectvalue=expectresultinfo[basename]
           if result_txt in expectvalue:
               result_value="TRUE"
           else:
               result_value = "FALSE"

           rows=[basename,expectvalue,result_txt,result_value]
           save(rows)
   print ('Message<<<<<<<<<<<<<<<<<<<<<<<<')

def on_error(ws, error):
    print ('error>>>>>>>>>>>>>>>>>>>>>>')
    print (error)
    print ('error<<<<<<<<<<<<<<<<<<<<<<')

def on_close(ws):
    print ("### closed ###")

def on_open(ws):
    def run(*args):
        step = 3000 #如果audioType是wav，此处需要修改为3200
        print(audio_file_path)
        with open(audio_file_path, 'rb') as f:
            while True:
                data = f.read(step)
                if data:
                    ws.send(data, ABNF.OPCODE_BINARY)
                if len(data) < step:
                    break
                time.sleep(0.1)

        ws.send(END_FLAG, ABNF.OPCODE_BINARY)
        time.sleep(0.6)
        ws.close()
    _thread.start_new_thread(run, ())

def getHandShakeParams():
    temp_data =base64.b64encode(bytes(param,'utf-8'))
    print(temp_data)
    paramBase64 = temp_data.decode(encoding='utf-8', errors='ignore')
    print(temp_data)
    curtime=str(time.time()).split(".")[0]
    print(curtime)
    signtype = "sha256";
    originStr = APIKEY + curtime + paramBase64;
    checksum =  sha256hex(originStr)
    handshakeParam = "?appid=" + APPID + "&checksum=" + checksum + "&curtime=" + curtime + "&param=" + paramBase64 + "&signtype=" + signtype;
    return handshakeParam

def sha256hex(data):
    sha256 = hashlib.sha256()
    sha256.update(data.encode())
    res = sha256.hexdigest()
    print("sha256加密结果:", res)
    return res


def get_small_list_tuple(all_case, thread_num):
    allcase_len = len(all_case)
    step_list = get_average_step_list(allcase_len, thread_num)
    if step_list==None:
        return tuple([all_case])
    else:
        print(step_list)
        out_list = []
        for i in range(thread_num):
            step = step_list[i]
            inner_list = all_case[:step]
            out_list.append(inner_list)
            for x in range(step_list[i]):
                del (all_case[0])
        return tuple(out_list)


def get_average_step_list(allcase_len, thread_num):
    if allcase_len <= 0 or 0 ==thread_num or allcase_len< thread_num:
         return None
    else:
        j = allcase_len / thread_num  # 除数,python除法不会四舍五入，52
        k = allcase_len % thread_num  # 余数要再被依次分配给n，直到分完,8
        step_list = []
        for i in range(allcase_len):
            # 先把除数分成n份，比如先分10分52
            step_list.append(int(j))
        if k > 0:  # 余数>0时, 除数再依次分配余数，直到分完
            for i in range(k):
                step_list[i] += 1
    return step_list

def gen_case(filelist,expect_resultlist,rootpath,child_resultpath,runlistinfo):
    global audio_file_path,expectresultinfo,save_resultpath
    expectresultinfo=expect_resultlist
    save_resultpath=child_resultpath
    print("文件大小:",str(len(filelist)))
    for one_file in filelist:
        if one_file not in expect_resultlist:
            print(one_file,"不存在")
            continue

        if one_file in runlistinfo:
            print("ddddddddddddddddddd")
            continue

        audio_file_path = rootpath + one_file
        print(audio_file_path)

        url = BASE_URL + getHandShakeParams()
        print(url)
        websocket.enableTrace(True)
        ws = websocket.WebSocketApp(url,header={"Origin":"http://wsapi.xfyun.cn"},
                                    on_message=on_message,
                                    on_error=on_error,
                                    on_close=on_close)
        ws.on_open = on_open
        ws.run_forever()


def run_process(path,dict_infovalue,child_resultpath):
    filelist = os.listdir(path)
    small_list_tuple = get_small_list_tuple(filelist, THREAD_NUM)
    processlist = []

    runlistinfo = read_fileinfo(child_resultpath)
    print("**********************************************")
    print(child_resultpath,runlistinfo)
    print("**********************************************")
    for i in range(len(small_list_tuple)):
        print(small_list_tuple[i])
        p_one = Process(target=gen_case, args=(small_list_tuple[i], dict_infovalue, path, child_resultpath,runlistinfo,))
        processlist.append(p_one)
        p_one.start()

    for i in processlist:
        i.join()

def create_path(path):
    if os.path.exists(path) == False:
        os.makedirs(path)
    return  path

def start_run_websocket(genlist,g_method):
    for key in genlist:
        keylist=genlist[key]
        currentpath=AUDIO_ROOT_PATH+keylist[0]
        current_resultpath=create_path(RESULT_ROOT_PATH+g_method+"/"+key+"/")
        is_exist_childdir=keylist[2]
        # 读取excel文件，并存储到dict中.
        excel_path=MANUAL_CASE_PATH+keylist[1]
        temp_resultinfo = returndictinfobyexcel(excel_path)
        if is_exist_childdir=="true":
            child_list = os.listdir(currentpath)
            for child_file in child_list:
                childpath=currentpath+child_file+"/"
                print(current_resultpath)
                child_resultpath=current_resultpath+child_file+".txt"
                run_process(childpath,temp_resultinfo,child_resultpath)
        else:
            run_process(currentpath, temp_resultinfo, current_resultpath+key+".txt")
    # 计算识别率
    #marge_excel(FILE_PATH)

def returndictinfobyexcel(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    allname = workbook.sheetnames
    dictinfo={}
    for one_sheet in allname:
        table = workbook[one_sheet]
        # 获取总行数
        nrows = table.max_row + 1
        for i in range(3, nrows):
            wav_value = table.cell(i,1).value  # wav文件值
            result_value = table.cell(i, 2).value  # 预期结果值
            dictinfo[wav_value]=result_value

    return dictinfo

if __name__ == "__main__":
    if os.path.exists(RESULT_ROOT_PATH)==False:
        os.makedirs(RESULT_ROOT_PATH)

    if GEN_METHOD=="all":
        pass
    else:
       genlist=RUN_LIST[GEN_METHOD]
       start_run_websocket(genlist,GEN_METHOD)



