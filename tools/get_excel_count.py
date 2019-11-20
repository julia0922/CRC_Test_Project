import  openpyxl
import os

'''
# excel 条数：209825
rootdir="D:/MyData/ex_shenjf1/Desktop/云知声/"
alist=os.listdir(rootdir)
allcount=0
for item in alist:
    excel_path=rootdir+item
    print(excel_path)
    if excel_path.endswith(".xlsx"):
        wb=openpyxl.load_workbook(excel_path)
        excelcount=0
        names = wb.get_sheet_names()
        for one_sheet in names:
            if one_sheet=="Sheet":
                continue

            ws=wb.get_sheet_by_name(one_sheet)
            excelcount=excelcount+(ws.max_row-2)
            allcount=allcount+(ws.max_row-2)

        print(item, excelcount)
print(allcount)
'''

rootdir="E:/sucess_addcmd/txt/"
alist=os.listdir(rootdir)
allcount=0
for item in alist:
    file_path = rootdir + item
    print(file_path)
    if file_path.endswith(".txt"):
        f = open(file_path, "r", encoding="utf-8")
        alllines = f.readlines()
        print(item,len(alllines))
        allcount=allcount+len(alllines)

print(allcount)
