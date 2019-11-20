#import pandas as pd

#data=pd.read_excel("D:/MyData/ex_shenjf1/Desktop/小雅星空（内容技能）测试case集-2.xlsx")
#a=data.drop_duplicates(subset=['Query'])
#a.to_excel("D:/MyData/ex_shenjf1/Desktop/小雅星空（内容技能）测试case集-3.xlsx")


import openpyxl
new_excel=openpyxl.Workbook()
new_sheet=new_excel.create_sheet("去重后的")

money_sheet=new_excel.create_sheet("付费内容")

money1_sheet=new_excel.create_sheet("无法为您播放")

excel=openpyxl.load_workbook("E:/小雅星空（内容技能）测试case集-1114.xlsx")

work_sheet=excel.get_sheet_by_name("Sheet1")

dict_info={}
for j in range(1, work_sheet.max_row + 1):
     query_value=work_sheet.cell(j, 1).value
     if query_value  not in dict_info:
         dict_info[query_value]=""

         response_txt=str(work_sheet.cell(j, 10).value)
         if response_txt.find("为您播放免费试听")!=-1:
             money_sheet.append([query_value, str(work_sheet.cell(j, 2).value),
                               str(work_sheet.cell(j, 3).value),
                               str(work_sheet.cell(j, 4).value),
                               str(work_sheet.cell(j, 5).value),
                               str(work_sheet.cell(j, 6).value),
                               str(work_sheet.cell(j, 7).value),
                               str(work_sheet.cell(j, 8).value),
                               str(work_sheet.cell(j, 9).value),
                               str(work_sheet.cell(j, 10).value)])
         elif response_txt.find("此设备暂时无法为您播放")!=-1:
             money1_sheet.append([query_value, str(work_sheet.cell(j, 2).value),
                                 str(work_sheet.cell(j, 3).value),
                                 str(work_sheet.cell(j, 4).value),
                                 str(work_sheet.cell(j, 5).value),
                                 str(work_sheet.cell(j, 6).value),
                                 str(work_sheet.cell(j, 7).value),
                                 str(work_sheet.cell(j, 8).value),
                                 str(work_sheet.cell(j, 9).value),
                                 str(work_sheet.cell(j, 10).value)])
         else:

             new_sheet.append([query_value,str(work_sheet.cell(j, 2).value),
                               str(work_sheet.cell(j, 3).value),
                               str(work_sheet.cell(j, 4).value),
                               str(work_sheet.cell(j, 5).value),
                               str(work_sheet.cell(j, 6).value),
                               str(work_sheet.cell(j, 7).value),
                               str(work_sheet.cell(j, 8).value),
                               str(work_sheet.cell(j, 9).value),
                               str(work_sheet.cell(j, 10).value)])


new_excel.save("E:/小雅os1.xlsx")


