# -*- coding:utf-8 -*-
import openpyxl

import os
from common.operation_file import OperationFile
from common.re_num_to_characters import ReplaceNumCharacters

excellist_dir="D:/audio_file/source音频/美的第一批递交/全品类/"
result_exceldir="D:/audio_file/wav音频/美的第一批递交/全品类.xlsx"
# 结果文件
result =openpyxl.Workbook()

# 获取表格列表

origin_file_list =OperationFile().getfilelistbydir(rootdir=excellist_dir,suffix=".xlsx")

print(origin_file_list)

rcobj=ReplaceNumCharacters()

# 循环遍历表格
for i in origin_file_list:
    print(i)

    file_path = i
    obj=openpyxl.load_workbook(file_path)
    allname=obj.get_sheet_names()
    print(allname)
    new_ws=result.create_sheet(os.path.basename(file_path).replace(".xlsx",""))
    new_ws.append(["Recognition success rate","","",""])
    new_ws.append(["wav","expect","asr","recognition result"])

    worksheet =obj.get_sheet_by_name(allname[0])
    rows = worksheet.max_row + 1
    print(rows)
    for j in range(2, rows):
        wav_value = str(worksheet.cell(j, 1).value).strip()

        if wav_value=="None":
            continue
        print(wav_value)
        expect_value = str(worksheet.cell(j, 2).value).strip()
        expect_value=rcobj.replace_num_value(expect_value)
        new_ws.append([wav_value,expect_value])

result.save(result_exceldir)
