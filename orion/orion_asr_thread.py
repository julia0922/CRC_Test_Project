import  os
import xlrd
import re
from xlutils.copy import copy
import  time
import  shutil

from threading import Thread
import openpyxl

# 需要用户指定的路径
# 猎户sit环境地址
orion_address_sit="wss://speech-test.ainirobot.com/ws/streaming-asr"

#音频用例文件路径
manual_rootdir="/data0/orion-tools/excel/"
#生成的结果路径
gen_rootdir="/data0/orion-tools/result/"
# 日志路径
log_rootdir="/data0/orion-tools/log/"

# 列表中的内容表示：
# 第一元素：音频文件的录制。
# 第二个元素，用例路径，统一放置到manual_rootdir目录下。
# 第三个元素：用两个选项，true和false。true表示是否包含sheet的名称，false表示不需要。
dict_info={"全品类17人":["/data1/audio_file/pcm/全品类17人语料/","全品类17人-测试语料-pcm.xlsx","true"],
           "家电百科":["/data1/audio_file/pcm/家电百科4人份/","4人份家电百科-测试案例.xlsx","true"],
           "故障码语料":["/data1/audio_file/pcm/20190828故障码录音/","故障码语料.xlsx","true"],
           "testin云测8月30号-空调": ["/data1/audio_file/pcm/testin云测8月30号递交数据/", "testin云测8月30号递交数据_空调.xlsx", "true"],
           "9月20日交付空调项目数据": ["/data1/audio_file/pcm/9月20日交付空调项目数据/", "9月20日交付空调项目数据.xlsx", "true"],
           "testin云测17人交付-空调": ["/data1/audio_file/pcm/testin云测第一批数据17人交付/", "testin云测第一批数据17人交付-空调.xlsx", "false"],}

# 线程数
threadNum=30

# 执行的方式： 有all，或者指定dict_info中的键值
gen_method="testin云测17人交付-空调"

def getavage(rows):
    total=rows
    j = total / threadNum  # 除数,python除法不会四舍五入，52
    q = total % threadNum

    # k = m % n  # 余数要再被依次分配给n，直到分完,8
    step_list = []
    for i in range(threadNum):
        step_list.append(int(j))

    if q > 0:  # 余数>0时, 除数再依次分配余数，直到分完
        for i in range(q):
            step_list[i] += 1

    return step_list

def main(root_audio_path,manual_path,log_path,is_containdir):
    # 打开excel所在的文件路径，获取excel内容
    data=openpyxl.load_workbook(manual_path)
    # 获取所有的names
    names = data.sheetnames
    for name in names:
        table = data.get_sheet_by_name(name)
        nrows = table.max_row
        print("总行数:" + str(nrows))

        # 计算一个线程下跑多少条记录
        l = []
        if nrows<threadNum:
            steplist=[nrows]
        else:
            steplist = getavage(nrows)
        print(steplist)

        containdir_value=name
        if is_containdir=="false":
            containdir_value=""

        for index in range(0,len(steplist)):
            if index==0:
                startindex=1
                endindex=steplist[index]
            else:
                newsteplist=steplist[:index]
                newsteplist2 = steplist[:index+1]
                startindex=sum(newsteplist)
                endindex=sum(newsteplist2)

            p = Thread(target=run,args=(startindex,endindex,table,containdir_value,root_audio_path,log_path))  # 多线程
            l.append(p)
            p.start()

        for p in l:
            p.join()

        data.save(manual_path)



def sum(list):
    sum_value=0
    for i in list:
        sum_value=sum_value+i
    return sum_value


def run(startindex,endindex,table,s_name,root_audio_path,log_path):
    print(startindex,endindex)
    for i in range(startindex+1, endindex+1):
        result_value =str(table.cell(i, 4).value)  # 预期结果值
        print("预期结果的值", result_value)
        
        if result_value.lower()=="true" or result_value.lower()=="false":
            continue

        print("***************************************")
        wav_value = str(table.cell(i, 1).value)  # pcm文件值
        if wav_value=="wav" or wav_value=="pcm":
            continue

        if wav_value.find(".pcm")==-1:
           wav_value=wav_value+".pcm"
        except_value =str(table.cell(i, 2).value)  # 获取期望值
        except_value = except_value.strip().replace("，", "").replace("？", "").replace(" ", "")
        print(wav_value)
        whole_path = root_audio_path +s_name +"/"+wav_value
        print(whole_path)
        whole_logpath = log_path +s_name  + wav_value.replace(".pcm", "") + ".txt"
        print(whole_logpath)

        commeline = "./qnet-test -server_url="+orion_address_sit+"  -compress_type=0  -pid=7016 -protocol=0 -audio=" + whole_path + " -output_file=" + whole_logpath + " -write_file_level=3"
        print(commeline)
        os.system(commeline)
        # 读取log日志文件
        asr_value = ""
        time.sleep(0.1)
        if os.path.exists(whole_logpath)==False:
            table.cell(i,5).value="没有生成whole_logpath文件"
            continue

        with open(whole_logpath, 'r', encoding='utf-8') as file_to_read:
            lines = file_to_read.read()  # 整行读取数据
            print(lines)
            print(lines)
            if lines != "" and lines != None:
                a = re.search("statistics asr text:(.*)", lines, flags=0)
                print("*****************",a)
                if a=="" or a==None:
                   continue

                asr_value = a.group(1)
                asr_value = str(asr_value).strip().replace(' ', '')
                print(asr_value)
                table.cell(i, 3).value = str(asr_value)
                table.cell(i,5).value=whole_logpath
                if (str(except_value).lower() == str(asr_value).lower()):
                    table.cell(i, 4).value = "TRUE"
                else:
                    table.cell(i, 4).value ="FALSE"
            else:
                table.cell(i, 5).value= "asr结果为空"

def marge_excel(save_manual_path):
    mywb = openpyxl.Workbook()
    new_sheet = mywb.create_sheet(index=0, title='sheet1')
    row1 = ["音频", "案例总数", "成功数", "失败数","识别率"]
    new_sheet.append(row1)
    count_rate_path=gen_rootdir+os.path.basename(save_manual_path).replace(".xlsx","")+"_rate.xlsx"

    data = openpyxl.load_workbook(save_manual_path)
    names = data.sheetnames
    for name in names:
        table = data.get_sheet_by_name(name)
        row2=[]
        row2.append(name)
        sum=0
        sucesscount=0
        failcount=0
        sucess_rate_value=0.0

        for row in range(3,table.max_row+1):
            result_value=str(table.cell(row, 4).value)
            sum=sum+1
            if result_value.lower()=="true" or result_value=="1":
                sucesscount=sucesscount+1
            else:
                failcount=failcount+1

        row2.append(sum)
        row2.append(sucesscount)
        row2.append(failcount)
        print(sum)
        print(sucesscount)
        print(failcount)
        if sum != 0:
            sucess_rate = (sucesscount / sum) * 100
            sucess_rate_value = "{:.2f}%".format(sucess_rate)
            row2.append(sucess_rate_value)
            table.cell(1,2).value=sucess_rate_value
        new_sheet.append(row2)
        data.save(save_manual_path)

    mywb.save(count_rate_path)

def genmanual(key_list,key_value):
    # 判断目录文件是否存在
    root_audio_path = key_list[0]
    old_manual_path = manual_rootdir + key_list[1]
    save_manual_path = gen_rootdir + key_list[1]
    if os.path.exists(root_audio_path) == False:
        print(root_audio_path + "文件不存在")

    if os.path.exists(old_manual_path) == False:
        print(old_manual_path + "用例文件不存在")
    else:
        if os.path.exists(save_manual_path) == False:
            shutil.copyfile(old_manual_path, save_manual_path)

    log_path = log_rootdir + key_value + "/"
    if os.path.exists(log_path) == False:
        os.makedirs(log_path)

    is_containdir = key_list[2]
    main(root_audio_path, save_manual_path, log_path, is_containdir)

    marge_excel(save_manual_path)  # 计算识别率

if __name__ == "__main__":
    if gen_method=="all":
       for key_value in dict_info:
           key_list = dict_info[key_value]
           genmanual(key_list,key_value)
    else:
        key_list = dict_info[gen_method]
        genmanual(key_list,gen_method)

