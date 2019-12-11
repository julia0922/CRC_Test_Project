#!/usr/bin/env python
# coding=utf-8
### websocket模块https://pypi.python.org/pypi/websocket-client
### wavfile是一个speex压缩的ogg文件,文件格式是Ogg data, Speex audio
import websocket
import re
from gevent import os
from websocket import ABNF
import time
import _thread
import json
from multiprocessing import Process,Manager
import  openpyxl
from common.operation_excel import OperationExcel
from common.common_service import Commonservice

# sit环境执行地址
product_id = '278575321' #填入产品Id
url="ws://linksit.aimidea.cn:10000/cloud/connect"

excel_path="E:/audiofile/100人粤语-采集-0910/粤语用例.xlsx"
#音频文件路径
root_cmd_path="E:/audiofile/100人粤语-采集-0910\结果数据/100人粤语-采集-批次2-0910/wav/"

# 线程数
threadNum=10
g_startindex=3  #开始行
log_path="log1.txt"

fileObject = open(log_path, 'w')

# 获取sheet
table =None


index=2
except_value=""  #预期结果

currentindex=0
success_total=0
fail_total=0
total=0


def on_message(ws, message):
   print ('Message>>>>>>>>>>>>>>>>>>>>>>>>')
   print(message)
   if(message == "音频为空"):
       text = "音频为空"
   else:
       mJson = json.loads(message)
       text=""
       if 'asr' in message:
           text = mJson['data']['asr']
           print(key_value,text)
           if key_value in return_data:
               print("ppppppppppppppppppppppppppppppppp")
               value_info=return_data[key_value]
               value_info["asr"]=text

               if text==value_info["expect"]:
                   value_info["result"] = "True"
               else:
                   value_info["result"] = "False"

               return_data[key_value]=value_info
               print(return_data)


   print('Message<<<<<<<<<<<<<<<<<<<<<<<<')
   '''

   global success_total,fail_total,total
   index=currentindex
   print(index)
   text=str(text).replace(' ', '').replace("，","")
   print("返回值:",text)
   pattern = re.compile(r'(。|\?|,|\.)')
   out_value = re.sub(pattern, '',text)
   table.cell(index, 3).value=out_value
   total=total+1
   except_value = table.cell(index, 2).value
   except_value = str(except_value).replace("，", "").replace("？", "").replace("（","").replace("）","")
   print("期望值：",except_value)
   if(except_value==out_value):
       table.cell(index, 4).value="True"
       success_total=success_total+1
   else:
       print("fail")
       print(index)
       table.cell(index, 4).value = "False"
       fail_total=fail_total+1

   print ('Message<<<<<<<<<<<<<<<<<<<<<<<<')
   '''

def on_error(ws, error):
    print ('error>>>>>>>>>>>>>>>>>>>>>>')
    print (error)
    save_sucessrate()
    print ('error<<<<<<<<<<<<<<<<<<<<<<')

def on_close(ws):
    print ("### closed ###")
    save_sucessrate()
    if isEnd==False:
        ws = websocket.WebSocketApp(url,
                                    on_message=on_message,
                                    on_error=on_error,
                                    on_close=on_close)
        ws.on_open = on_open
        ws.run_forever()

def save_sucessrate():
    if total != 0:
        sucess_rate = (success_total / total) * 100
        sucess_rate_value = "{:.2f}%".format(sucess_rate)
        table.cell(1, 2).value=sucess_rate_value
        print(wb)
        wb.save(excel_path)



def on_open(ws):
    def run(*args):
        global currentindex, table
        global success_total, fail_total, total
        global isEnd
        global key_value
        isEnd=False

        # 获取数据
        for one_dict in list_info:
            print(one_dict)
            for key, value in one_dict.items():
                key_value=key
                wav_path=audiofile_path+"/"+ key
                if (os.path.exists(wav_path) == False):
                    print("音频路径为空")
                    continue

                content = {
                    "topic": "cloud.speech.trans",
                    "mid": "1508232047194123",
                    "version": "1.0",
                    "request": {
                        "timestamp": 1508232047199,
                        "sessionId": "aaaadsfasdffsdf123"
                    },
                    "params": {
                        "audio": {
                            "audioType": "wav",
                            "sampleRate": 16000,
                            "channel": 1,
                            "sampleBytes": 2
                        },
                        "asrIsp": "xfyun",
                        "ttsIsp": "xfyun"
                    }
                }
                opencontent = {
                    "topic": "cloud.connect",
                    "mid": "1508232047154",
                    "version": "1.0",
                    "request": {
                        "apiVer": "1.0.0",
                        "timestamp": 1508232047194,
                        "pki": "aaaadsfasdffqwe"
                    },
                    "params": {
                        "sn": "0000DB11138104887174101101744561",
                        "category": "B1",
                        "model": "456",
                        "id": "1099511840456",
                        "ip": "0.0.0.0",
                        "mac": "88e9fe5d3456",
                        "random": "545123"
                    }
                }
                ws.send(json.dumps(opencontent), ABNF.OPCODE_TEXT)
                ws.send(json.dumps(content), ABNF.OPCODE_TEXT)
                step = 3200  # 如果audioType是wav，此处需要修改为3200
                with open(wav_path, 'rb') as f:
                    while True:
                        data = f.read(step)
                        if data:
                            ws.send(data, ABNF.OPCODE_BINARY)
                        if len(data) < step:
                            break
                        time.sleep(0.1)
                ws.send('', ABNF.OPCODE_BINARY)
                time.sleep(2)

        print("OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO")
        print(return_data)
        '''
        # 获取所有的names
        names = wb.sheetnames
        for name in names:
            # 获取sheet
            success_total = 0
            fail_total = 0
            total = 0

            print("总案例数:" + str(total))
            table = wb.get_sheet_by_name(name)
            print("当前sheet：" + name)
            # 获取总行数
            nrows = table.max_row+1
            print("总行数:" + str(nrows))
            cmd_path = root_cmd_path+name
            print(cmd_path)
            currentindex = 0
            for i in range(index, nrows):
                wav_value = table.cell(i, 1).value  # wav文件值
                if wav_value.strip() == '':
                    continue
                print("path:" + wav_value)
                result_value =str(table.cell(i, 4).value)  # 预期结果值
                print(result_value)
                if result_value.strip()!="" and result_value!="None":
                    continue

                ret1 = wav_value.find(".wav")
                if ret1 == -1:
                    wav_value=wav_value+".wav"

                wav_path = cmd_path + "/" + wav_value  # wav 路径
                print("***************************************")
                print(wav_path)
                currentindex = i
                print("当前index" + str(currentindex))
                if (os.path.exists(wav_path) == False):
                    print("音频路径为空")
                    continue
                print("index:"+str(i))
                content = {
                    "topic": "cloud.speech.trans",
                    "mid": "1508232047194123",
                    "version": "1.0",
                    "request": {
                        "timestamp": 1508232047199,
                        "sessionId": "aaaadsfasdffsdf123"
                    },
                    "params": {
                        "audio": {
                            "audioType": "wav",
                            "sampleRate": 16000,
                            "channel": 1,
                            "sampleBytes": 2
                        },
                        "asrIsp": "xfyun",
                        "ttsIsp": "xfyun"
                    }
                }
                opencontent = {
                    "topic": "cloud.connect",
                    "mid": "1508232047154",
                    "version": "1.0",
                    "request": {
                        "apiVer": "1.0.0",
                        "timestamp": 1508232047194,
                        "pki": "aaaadsfasdffqwe"
                    },
                    "params": {
                        "sn": "0000DB11138104887174101101744561",
                        "category": "B1",
                        "model": "456",
                        "id": "1099511840456",
                        "ip": "0.0.0.0",
                        "mac": "88e9fe5d3456",
                        "random": "545123"
                    }
                }
                ws.send(json.dumps(opencontent), ABNF.OPCODE_TEXT)
                ws.send(json.dumps(content), ABNF.OPCODE_TEXT)
                step = 3200  # 如果audioType是wav，此处需要修改为3200
                with open(wav_path, 'rb') as f:
                    while True:
                        data = f.read(step)
                        if data:
                            ws.send(data, ABNF.OPCODE_BINARY)
                        if len(data) < step:
                            break
                        time.sleep(0.1)
                ws.send('', ABNF.OPCODE_BINARY)
                time.sleep(2)
            save_sucessrate()
        '''
        isEnd=True
        ws.close()
    _thread.start_new_thread(run, ())


def start_single_process(d,list,audiopath):
    global return_data,list_info,audiofile_path
    return_data = d
    list_info=list
    audiofile_path=audiopath

    '''
    valueinfo = return_data["039M23_F01_41_0491.wav"]
    valueinfo['asr'] = "dddddd"
    valueinfo['result'] = "False"
    return_data["039M23_F01_41_0491.wav"] = valueinfo

    # print(return_data)
    print("**************************")
    print(return_data)

    '''
    print(return_data)
    print("***********************************************")
    websocket.enableTrace(True)
    ws = websocket.WebSocketApp(url,
                                on_message=on_message,
                                on_error=on_error,
                                on_close=on_close)
    ws.on_open = on_open
    ws.run_forever()


def get_processlist(dict_info,steplist,cmd_path):
    processlist=[]
    for index in range(0, len(steplist)):
        if index == 0:
            startindex = 0
            endindex = steplist[index]
        else:
            newsteplist = steplist[:index]
            newsteplist2 = steplist[:index + 1]
            startindex = sum(newsteplist)
            endindex = sum(newsteplist2)

        temp_data = list[startindex:endindex]
        p_one = Process(target=start_single_process, args=(temp_data, dict_info,cmd_path,))
        processlist.append(p_one)
        p_one.start()
    return processlist




if __name__ == "__main__":
    excelobj = OperationExcel(excel_path)
    cs = Commonservice()
    workbook_obj, names = excelobj.get_all_sheetnames()  # 获取所有的sheet
    print(names)
    for name in names:
        print(name)
        table_obj = workbook_obj.get_sheet_by_name(name)  # 根据pandas的方式读取excel文件
        # 多个进程共享dict_info的数据
        temp_dict_info, list = cs.getdict_byexcelsheet(g_startindex, table_obj)

        steplist = cs.getavage(len(list), threadNum)  # 根据进程数分发，一个进行执行多少个wav文件。
        cmd_path = root_cmd_path + name + "/"
        processlist = []
        for index in range(0, len(steplist)):
            if index == 0:
                startindex = 0
                endindex = steplist[index]
            else:
                newsteplist = steplist[:index]
                newsteplist2 = steplist[:index + 1]
                startindex = sum(newsteplist)
                endindex = sum(newsteplist2)

            temp_data = list[startindex:endindex]
            p_one = Process(target=start_single_process, args=(temp_dict_info,list,cmd_path))
            p_one.start()
            processlist.append(p_one)


        for i in processlist:
            i.join()

        print(temp_dict_info)

        for i in range(3, table_obj.max_row + 1):
            wav_value = str(table_obj.cell(i, 1).value)
            print(wav_value)
            if wav_value in temp_dict_info:
                wav_info=temp_dict_info[wav_value]
                table_obj.cell(i, 3).value=str(wav_info['asr'])
                table_obj.cell(i,4).value = str(wav_info['result'])

    workbook_obj.save(excel_path)



    print("执行完成")
