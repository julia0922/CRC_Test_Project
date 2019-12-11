import os,shutil
import openpyxl
from openpyxl.styles import Font,NamedStyle,Border,Side
import xlrd
rootPath="E:/audiofile/100人粤语-采集-0910/"

workbook=openpyxl.Workbook()

def read_excel():
    excel="E:/audiofile/录制音频预期文本/3.方言相关录音需求(0904).xlsx"
    b=openpyxl.load_workbook(excel)
    s=b.get_sheet_by_name("广东话词表-已转写")
    hashlist={}
    for i in range(1, s.max_row + 1):
        text = str(s.cell(i, 2).value)  #语料
        id=str(s.cell(i, 4).value)   # id
        print(id)
        if id not in hashlist.keys():
            hashlist[id]=text

    return hashlist



def marge_excel(h):
    list_dir=os.listdir(rootPath)
    for excel_file in list_dir:
        excel_file_path=rootPath+excel_file
        if os.path.isfile(excel_file_path) and excel_file.find(".xlsx")!=-1:
            print(excel_file_path)
            tempfilename= excel_file_path.split("/")
            filename= str(tempfilename[len(tempfilename)-1]).replace(".xlsx","")
            f = filename.split("-")
            f_name = f[len(f) - 1]
            print(f_name)
            n_sheet=workbook.create_sheet(f_name)
            b=openpyxl.load_workbook(excel_file_path)
            w_sheet=b.get_sheet_by_name("Sheet1")
            row=1
            for i in range(1, w_sheet.max_row + 1):
                wav_value = str(w_sheet.cell(i, 2).value)
                a = wav_value.split("_")
                m_value=str(a[2])
                print(wav_value)
                if m_value=="14" or m_value=="41":
                    n_sheet.cell(row,1).value=wav_value
                    text_value = str(a[2]) + "_" + str(a[3]).replace(".wav","")
                    if text_value in h.keys():
                        n_sheet.cell(row,2).value = h[text_value]

                    n_sheet.cell(row, 3).value = str(w_sheet.cell(i, 3).value)
                    row=row+1

    workbook.save("E:/audiofile/100人粤语-采集-0910/用例.xlsx")



if __name__=="__main__":
    #wav_value="036F35_F01_04_0073.wav"
    #a=wav_value.split("_")
    #print(a)
    #text_value = str(a[2]) + "_" + str(a[3]).replace(".wav","")
    #print(text_value)
    #excel_file_path="E:/audiofile/100人粤语-采集-0910/100人粤语-采集-批次2-0910nameinfo-G00036.xlsx"
    #tempfilename = excel_file_path.split("/")
    #filename = str(tempfilename[len(tempfilename) - 1]).replace(".xlsx", "")
    #f = filename.split("-")
    #print(f)
    #f_name = f[len(f) - 1]
    #print(f_name)

    h=read_excel()
    print(h)
    marge_excel(h)
