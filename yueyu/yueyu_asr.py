#!/usr/bin/env python
# coding=utf-8
### websocket模块https://pypi.python.org/pypi/websocket-client
### wavfile是一个speex压缩的ogg文件,文件格式是Ogg data, Speex audio
import websocket
import re
from gevent import os
from websocket import ABNF
import time
import _thread
import json
import xlrd
from xlutils.copy import copy
import  openpyxl

# sit环境执行地址
product_id = '278575321' #填入产品Id
url="ws://linksit.aimidea.cn:10000/cloud/connect"

excel_path="E:/audiofile/100人粤语-采集-0910/用例.xlsx"

#音频文件路径
root_cmd_path="E:/audiofile/100人粤语-采集-0910\结果数据/100人粤语-采集-批次2-0910/wav/"

log_path="log1.txt"

fileObject = open(log_path, 'w')

# 获取sheet
table =None


index=2
except_value=""  #预期结果

currentindex=0
success_total=0
fail_total=0
total=0
wb = openpyxl.load_workbook(excel_path)

def on_message(ws, message):
   print ('Message>>>>>>>>>>>>>>>>>>>>>>>>')
   print(message)
   if(message == "音频为空"):
       text = "音频为空"
   else:
       mJson = json.loads(message)
       text=""
       if 'asr' in message:
         text = mJson['data']['asr']

   global success_total,fail_total,total
   index=currentindex
   print(index)
   text=str(text).replace(' ', '').replace("，","")
   print("返回值:",text)
   pattern = re.compile(r'(。|\?|,|\.)')
   out_value = re.sub(pattern, '',text)
   table.cell(index, 3).value=out_value
   total=total+1
   except_value = table.cell(index, 2).value
   except_value = str(except_value).replace("，", "").replace("？", "").replace("（","").replace("）","")
   print("期望值：",except_value)
   if(except_value==out_value):
       table.cell(index, 4).value="True"
       success_total=success_total+1
   else:
       print("fail")
       print(index)
       table.cell(index, 4).value = "False"
       fail_total=fail_total+1

   print ('Message<<<<<<<<<<<<<<<<<<<<<<<<')

def on_error(ws, error):
    print ('error>>>>>>>>>>>>>>>>>>>>>>')
    print (error)
    save_sucessrate()
    print ('error<<<<<<<<<<<<<<<<<<<<<<')

def on_close(ws):
    print ("### closed ###")
    save_sucessrate()
    if isEnd==False:
        ws = websocket.WebSocketApp(url,
                                    on_message=on_message,
                                    on_error=on_error,
                                    on_close=on_close)
        ws.on_open = on_open
        ws.run_forever()

def save_sucessrate():
    if total != 0:
        sucess_rate = (success_total / total) * 100
        sucess_rate_value = "{:.2f}%".format(sucess_rate)
        table.cell(1, 2).value=sucess_rate_value
        print(wb)
        wb.save(excel_path)



def on_open(ws):
    def run(*args):
        global currentindex, table
        global success_total, fail_total, total
        global isEnd
        isEnd=False

        # 获取数据

        # 获取所有的names
        names = wb.sheetnames
        for name in names:
            # 获取sheet
            success_total = 0
            fail_total = 0
            total = 0

            print("总案例数:" + str(total))
            table = wb.get_sheet_by_name(name)
            print("当前sheet：" + name)
            # 获取总行数
            nrows = table.max_row+1
            print("总行数:" + str(nrows))
            cmd_path = root_cmd_path+name
            print(cmd_path)
            currentindex = 0
            for i in range(index, nrows):
                wav_value = table.cell(i, 1).value  # wav文件值
                if wav_value.strip() == '':
                    continue
                print("path:" + wav_value)
                result_value =str(table.cell(i, 4).value)  # 预期结果值
                print(result_value)
                if result_value.strip()!="" and result_value!="None":
                    continue

                ret1 = wav_value.find(".wav")
                if ret1 == -1:
                    wav_value=wav_value+".wav"

                wav_path = cmd_path + "/" + wav_value  # wav 路径
                print("***************************************")
                print(wav_path)
                currentindex = i
                print("当前index" + str(currentindex))
                if (os.path.exists(wav_path) == False):
                    print("音频路径为空")
                    continue
                print("index:"+str(i))
                content = {
                    "topic": "cloud.speech.trans",
                    "mid": "1508232047194123",
                    "version": "1.0",
                    "request": {
                        "timestamp": 1508232047199,
                        "sessionId": "aaaadsfasdffsdf123"
                    },
                    "params": {
                        "audio": {
                            "audioType": "wav",
                            "sampleRate": 16000,
                            "channel": 1,
                            "sampleBytes": 2
                        },
                        "asrIsp": "xfyun",
                        "ttsIsp": "xfyun"
                    }
                }
                opencontent = {
                    "topic": "cloud.connect",
                    "mid": "1508232047154",
                    "version": "1.0",
                    "request": {
                        "apiVer": "1.0.0",
                        "timestamp": 1508232047194,
                        "pki": "aaaadsfasdffqwe"
                    },
                    "params": {
                        "sn": "0000DB11138104887174101101744561",
                        "category": "B1",
                        "model": "456",
                        "id": "1099511840456",
                        "ip": "0.0.0.0",
                        "mac": "88e9fe5d3456",
                        "random": "545123"
                    }
                }
                ws.send(json.dumps(opencontent), ABNF.OPCODE_TEXT)
                ws.send(json.dumps(content), ABNF.OPCODE_TEXT)
                step = 3200  # 如果audioType是wav，此处需要修改为3200
                with open(wav_path, 'rb') as f:
                    while True:
                        data = f.read(step)
                        if data:
                            ws.send(data, ABNF.OPCODE_BINARY)
                        if len(data) < step:
                            break
                        time.sleep(0.1)
                ws.send('', ABNF.OPCODE_BINARY)
                time.sleep(2)
            save_sucessrate()
        isEnd=True
        ws.close()
    _thread.start_new_thread(run, ())

if __name__ == "__main__":
    websocket.enableTrace(True)
    ws = websocket.WebSocketApp(url,
                              on_message = on_message,
                              on_error = on_error,
                              on_close = on_close)
    ws.on_open = on_open
    ws.run_forever()
