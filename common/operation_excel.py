import os
import pandas as pd
import openpyxl
from common.operation_file import OperationFile


class OperationExcel(object):
    def __init__(self,excelpath=None):
        if excelpath==None:
            self.path=None
        else:
            self.path=excelpath  # 文件名称
            if os.path.isfile(self.path) == False:
                self.isExcel=False
            elif os.path.exists(self.path) == False:
                self.isExcel = False
            else:
                self.isExcel = True

    # 通过pandas的库读取excel文件
    def read_excel_bypd(self,excelName=None):
        if self.isExcel==False:
            return None

        if excelName==None:
            excelobj = pd.read_excel(self.path, skiprows=1, skipfooter=1)
        else:
            excelobj=pd.read_excel(self.path,skiprows=1, skipfooter=1,sheet_name=excelName)

        return excelobj

    # 获取excel对象
    def getworkbook_obj(self):
        if self.isExcel == False:
            return None
        workbook = openpyxl.load_workbook(self.path)
        return workbook

    # 获取所有excel的sheet信息
    def get_all_sheetnames(self):
        obj=self.getworkbook_obj()
        print(obj)
        return obj,obj.get_sheet_names()

    # 根据结果信息获取信息
    def get_wavinfobyresult(self,obj=None,excelName=None,result_condition=False):
        if obj==None:
            obj=self.getworkbook_obj()

        list=[]
        if excelName!=None:
            obj_sheet=obj.get_sheet_by_name(excelName)
            for j in range(1, obj_sheet.max_row + 1):
                result_value = obj_sheet.cell(j, 4).value   # 预期结果的值
                wav_value = obj_sheet.cell(j, 1).value    # wav的信息
                if result_value==str(result_condition):
                    list.append(wav_value)
        return list


    # 根据index获取值,默认index为1
    def getcell_value(self,obj=None,excelName=None,index=1):
        if obj==None:
            obj=self.getworkbook_obj()
        list=[]
        if excelName!=None:
            obj_sheet=obj.get_sheet_by_name(excelName)
            for j in range(3, obj_sheet.max_row + 1):
                wav_value = obj_sheet.cell(j, index).value    # wav的信息
                if wav_value=="":
                    continue

                if str(wav_value) in list:
                    continue
                else:
                    list.append(wav_value)

        return list


    # 清除用例信息
    def clear_excel_info(self, obj=None, excelName=None):
        if obj == None:
            obj = self.getworkbook_obj()
        list = []
        if excelName != None:
            obj_sheet = obj.get_sheet_by_name(excelName)
            obj_sheet.cell(1, 2).value = ""
            for j in range(3, obj_sheet.max_row + 1):
                obj_sheet.cell(j, 3).value=""
                obj_sheet.cell(j, 4).value = ""
                obj_sheet.cell(j, 5).value = ""
        else:
            names=obj.sheetnames
            for item in names:
                obj_sheet = obj.get_sheet_by_name(item)
                obj_sheet.cell(1,2).value=""
                for j in range(3, obj_sheet.max_row + 1):
                    obj_sheet.cell(j, 3).value = ""
                    obj_sheet.cell(j, 4).value = ""
                    obj_sheet.cell(j, 5).value = ""

        obj.save(self.path)

    # 获取映射表中的信息，并返回dict格式的信息
    def get_dict_info(self,gen_type):
        if gen_type==1:
            excelpath="2.空调相关录音需求.xlsx"
            sheet_name="修改后的成人普通话词表"
        elif gen_type==2:
            excelpath="4.全品类相关录音需求.xlsx"
            sheet_name="指令词"

        abs_path=(os.path.abspath('..'))+"/audiofile/"+excelpath
        print(abs_path)
        workbook = openpyxl.load_workbook(abs_path)
        w_sheet = workbook.get_sheet_by_name(sheet_name)
        rows = w_sheet.max_row + 1
        hashtable_value = {}
        for ind in range(1, rows):
            key_item = w_sheet.cell(ind, 3).value
            key_value = w_sheet.cell(ind, 2).value
            print(key_item, key_value)
            hashtable_value[key_item] = key_value
        return hashtable_value

    # 写excel用例文件
    def write_excel(self,workbook,audiofile_rootdir,gen_type):
        dict_info = self.get_dict_info(gen_type)
        file_list=OperationFile().getfilelistbydir(audiofile_rootdir)
        titlevalue=audiofile_rootdir.split("/")[-2]
        worksheet = workbook.create_sheet(title=titlevalue)
        worksheet.append(["Recognition success rate", ""])
        worksheet.append(["wav", "expect", "asr", "recognition result"])

        for item in file_list:
            dirname_list = item.replace(".wav", "").split("_")
            print(dirname_list)
            suffixvalue = dirname_list[2] + "_" + dirname_list[3]
            exceptvalue=""
            if suffixvalue in dict_info:
                exceptvalue = dict_info[suffixvalue]

            worksheet.append([item,exceptvalue])

        workbook.save(self.path)

    # 移除掉指定的sheet的信息
    def  remove_sheetbyname(self,sheetname):
        workbook = openpyxl.load_workbook(self.path)
        names = workbook.get_sheet_names()
        print(names)
        if sheetname in names:
            workbook.remove(workbook.get_sheet_by_name(sheetname))
        workbook.save(self.path)

    def write_ratetoexcel_bydict(self,dictlist):
        workbook=openpyxl.Workbook()
        worksheet = workbook.create_sheet(title="识别率")
        worksheet.append(["语料","总数","成功","失败","成功率"])
        for item in dictlist:
            info=dictlist[item]
            if isinstance(info, list):
                info.insert(0,item)
                worksheet.append(info)
            elif isinstance(info,dict):
                person = info['person']
                pass_value = info['pass']
                sucess_rate = (pass_value / person) * 100
                sucess_rate_value = "{:.2f}%".format(sucess_rate)
                worksheet.append([item, person, pass_value, info['fail'], sucess_rate_value])
        workbook.save(self.path)

    # 获取excel文件中每个sheet的识别率情况，返回一个dict
    def getsuccessrate(self):
        rate_dict = {}
        obj,allname=self.get_all_sheetnames()
        for item in allname:
            worksheet=obj.get_sheet_by_name(item)
            rows = worksheet.max_row + 1
            sum=failcount=sucesscount=0
            onepersionlist=[]

            for ind in range(3, rows):
                result_value = str(worksheet.cell(ind, 4).value)
                sum = sum + 1
                if result_value.lower() == "true" or result_value == "1":
                    sucesscount = sucesscount + 1
                else:
                    failcount = failcount + 1

            if sum != 0:
                sucess_rate = (sucesscount / sum) * 100
                sucess_rate_value = "{:.2f}%".format(sucess_rate)
                worksheet.cell(1,2).value=sucess_rate_value

            onepersionlist.append(sum)
            onepersionlist.append(sucesscount)
            onepersionlist.append(failcount)
            onepersionlist.append(sucess_rate_value)

            rate_dict[item]=onepersionlist
        obj.save(self.path)
        return rate_dict

    def get_common_rate(self):
        obj, allname = self.get_all_sheetnames()
        common_rate_dict={}
        for item in allname:
            worksheet = obj.get_sheet_by_name(item)
            rows = worksheet.max_row + 1
            for j in range(3, rows):
                except_value = str(worksheet.cell(j, 2).value).strip()
                result_value = str(worksheet.cell(j, 4).value).strip()
                if except_value not in common_rate_dict.keys():
                    result_dict = {"person":1,"pass":0,"fail":0}
                    if result_value.lower() == "true" or result_value == "1":
                        result_dict["pass"] = 1
                    else:
                        result_dict["fail"] = 1
                    common_rate_dict[except_value] = result_dict
                else:
                    result_dict = common_rate_dict[except_value]  # 获取出result_dict的信息
                    if result_value.lower() == "true" or result_value == "1":
                        result_dict["pass"] = result_dict["pass"] + 1
                    else:
                        result_dict["fail"] = result_dict["fail"] + 1
                    result_dict["person"] = result_dict["person"] + 1
                    common_rate_dict[except_value] = result_dict

        return common_rate_dict








