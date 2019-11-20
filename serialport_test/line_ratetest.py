# -*- coding: UTF-8 -*-
import os
import winsound
import time
import serial
import serial.tools.list_ports
import shutil
import openpyxl

'''
串口端打印ASR
脚本控制本地唤醒词音频播放
——判断是否唤醒成功
——唤醒成功，则播放音频语料（如果没有唤醒成功，则再次播放唤醒音频，直到唤醒成功为止）
——读取串口端打印的ASR结果
——将此ASR结果和期望值比较
——完全一样，pass。不一样，fail。
需要注意的是：
1.要统计播放唤醒音频测试；
2.需要统计唤醒成功次数；
3.需要打印唤醒成功率；
4.需要打印识别成功率。

作者：
创建时间：
'''

# 全局变量-文件配置
wakeup_path = "D:/aicolud_file/001M26_01_40_0001.wav"   # #唤醒文件：你好小美的音频文件
audio_root_path =  "E:/audiofile/离在线音频/（非离线）纯在线本机&场景指令词表/"  #需要执行的音频文件路径
excel_path = 'E:/audiofile/离在线音频/（非离线）纯在线本机&场景指令词表.xlsx'   # 映射关系表，命令词
save_excelpath="E:/audiofile/离在线音频/测试结果.xlsx"   # 测试结果保存的文件路径
file_copy_rpath="E:/audiofile/命令的文件/"   # 拷贝路径.

# 查询可用串口
def query():
    plist = list(serial.tools.list_ports.comports())
    if len(plist) <= 0:
        print("The Serial port can't find!")
        return None
    else:
        #plist_0 =list(plist[0])
        #serialName = plist_0[0]

        #串口地址配置
        serialName="COM3"
        serialFd = serial.Serial(serialName, 9600, timeout=60)
        #serialFd = serial.Serial(serialName,921600,timeout = 60)
        print("check which port was really used > ",serialFd.name)

    return serialFd

# 打开串口
def open_interface(serialFd):
    if serialFd.isOpen():
        print("Serial port open success")
    else:
        print("Serial port open failed")

# 定义识别率 打断成功率 唤醒信息
def clear_data():
    # 识别信息
    global  cntCmdTotal,cntCmdSucc,cntCmdFail,ratioCmd
    cntCmdTotal = 0 # 案例总数
    cntCmdSucc = 0  # 识别成功次数
    cntCmdFail = 0  # 识别失败次数
    ratioCmd = 0.0  # 识别率

    # 唤醒信息
    global totalWk, cntWkSucc, cntWkFail,ratioWk
    totalWk = 0    # 总唤醒次数
    cntWkSucc = 0  # 唤醒成功次数
    cntWkFail = 0  # 唤醒失败次数
    ratioWk = 0.0  # 唤醒成功率

def returndictinfobyexcel(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    allname = workbook.get_sheet_names()
    dictinfo={}
    for one_sheet in allname:
        table = workbook.get_sheet_by_name(one_sheet)
        # 获取总行数
        nrows = table.max_row + 1
        for i in range(1, nrows):
            wav_value = table.cell(i,1).value  # wav文件值
            result_value = table.cell(i, 2).value  # 预期结果值
            dictinfo[wav_value]=result_value

    return  dictinfo

 # 拷贝某一个文件
def start_copy_file(s_path,taget_path):
     if os.path.exists(s_path) == True:
         root_dirpath = os.path.dirname(taget_path)
         print(root_dirpath)
         if os.path.exists(root_dirpath)==False:
             os.makedirs(root_dirpath)
         shutil.copyfile(s_path, taget_path)

# 获取excel，计算识别率 打断成功率 唤醒信息
def start_play_audio(serialFd,dictinfo):
    new_wb=openpyxl.Workbook();
    new_ws = new_wb.create_sheet(title="命中率")
    new_ws.append(["在线语料",'命中本地ASR','命中置信度','命中的录音文件名称'])
    list_dir=os.listdir(audio_root_path)
    for one_dir in list_dir:
        childpath=audio_root_path+one_dir
        list_child_dir=os.listdir(childpath)
        hitcounts=0  # 命中次数
        miss_hitcounts=0  # 未命中次数
        for one_file in list_child_dir:
            child_filepath=childpath+"/"+one_file
            if (os.path.exists(child_filepath) == False):
                print("文件不存在")
                continue

            issucess = wakeup_File(serialFd)  # 唤醒小美
            if issucess:
                # step 2-2, 播放命令词，读取串口，判断是否正确识别
                time.sleep(1)  # 间隔1秒.
                winsound.PlaySound(child_filepath, winsound.SND_FILENAME)
                ars_result = recvCmd(serialFd)  # 接收asr返回结果
                print("asr返回结果信息:" + ars_result)
                asrinfo_value=""  # asr的信息
                conf_value=""  # 置信度信息
                if ars_result!=""  and  ars_result.find("asr")!=-1:
                    hitcounts=hitcounts+1
                    asr_list=ars_result.split(":")
                    asrinfo_value=asr_list[1].replace("\\conf","").replace("]","").strip()
                    conf_value=asr_list[2].replace("]","").strip()
                    print(asrinfo_value,conf_value)

                    # 拷贝文件
                    #copy_tagetpath=file_copy_rpath+one_dir+"/"+one_file
                    #start_copy_file(child_filepath,copy_tagetpath)

                else:
                    miss_hitcounts=miss_hitcounts+1

                if dictinfo[one_file]!=None:
                    except_info=str(dictinfo[one_file]).strip()
                    print(except_info)

                new_ws.append([except_info,asrinfo_value,conf_value,child_filepath])

        totalcount=len(list_child_dir)
        ratioWk = (hitcounts / totalcount) * 100
        rate_value = "{:.2f}%".format(ratioWk)
        str_totalcount="总次数:"+str(totalcount)
        str_hitcounts="命中次数:"+ str(hitcounts)
        str_miss_hitcounts="未命中次数:"+str(miss_hitcounts)
        str_rate="命中率:" + str(rate_value)
        print(str_totalcount)
        print(str_hitcounts)
        print(str_miss_hitcounts)
        print(str_rate)

        print("命中率:" , rate_value)
        new_ws.append([str_totalcount, str_hitcounts, str_miss_hitcounts,str_rate])
        new_wb.save(save_excelpath)

def recvCmd(serial):
    global cntCmdSucc,cntCmdFail,cntCmdTotal
    i = 0
    res = 'ASR结果 '
    # 10 秒内收到回应
    cmd_newData=""
    try:
        while i < 20:
            count = serial.inWaiting()
            print("count: " + str(count))
            if count > 0:
                time.sleep(0.1)
                buff = serial.read_all()
                print("buff: " + str(buff))
                if isinstance(buff, bytes):
                    temp_data = buff.decode(encoding='utf-8', errors='ignore')
                    cmd_newData = cmd_newData + temp_data
                    break
            else:
                time.sleep(0.2)
                i += 1
    except Exception as e:
           #fileObject.write("获取asr结果异常："+str(e))
           print("获取asr结果异常："+str(e))

    if len(cmd_newData) <= 1:
        cmd_newData = ''

    #print("ASR结果：" + cmd_newData)
    cmd_newData=cmd_newData.replace(" ","")
    serial.flushInput()
    return cmd_newData


# 唤醒小美,读取音频文件
def  wakeup_File(serialFd):
    iswake = False
    istts=False
    winsound.PlaySound(wakeup_path, winsound.SND_FILENAME)  # 播放唤醒词
    is_wakeup,data= recvWakeup(serialFd)    # 接收是否唤醒成功
    #是否唤醒
    print("是否唤醒: " + str(is_wakeup))
    if is_wakeup == True:
         iswake=True
    else:
        iswake=wakeup_File(serialFd)
    return iswake

#唤醒小美,获取唤醒
def recvWakeup(serial):
    try:
        txt = 'xiaomei'
        i = 0
        new_data = ""
        while(i < 50):
            count = serial.inWaiting()
            print("count: " + str(count))
            if count > 0:
                buff = serial.read(count)
                print(type(buff))
                print("buff: " + str(buff))
                if isinstance(buff, bytes):
                    temp_data=buff.decode(encoding='utf-8',errors='ignore')
                    new_data =new_data+ temp_data
                    print("测试数据："+new_data)
                break
            else:
                time.sleep(0.1)
                i += 1

        ret = new_data.find(txt)
        serial.flushInput()
        if ret != -1:
            print('……唤醒成功……')
            return True,new_data
        else:
            print('……唤醒失败……')
            return False,new_data
    except Exception as e:
        #fileObject.write("获取唤醒语句异常: " + str(e))
        print("异常信息如下: "+str(e))
        return False,""

if __name__=="__main__":
    serialFd=query()   # 查询是否有插入串口
    if serialFd!=None:
        open_interface(serialFd)  # 打开串口
        if (os.path.exists(wakeup_path)==False):
            print("唤醒的音频文件不存在: " + wakeup_path)
        else:
            #读取excel文件，并存储到dict中.
            dict_infovalue=returndictinfobyexcel(excel_path)
            start_play_audio(serialFd,dict_infovalue)







