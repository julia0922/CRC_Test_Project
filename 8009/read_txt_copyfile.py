import  openpyxl
import  os
import  shutil
from common.operation_file import OperationFile

copy_dir="C:/8009 音频检查/陈工/"
source_dir="C:/8009 音频检查/"
txt_path="C:/8009 音频检查/陈工.txt"

new_txt_path="C:/8009 音频检查/new_陈工.txt"

excel_path="C:/8009 音频检查/3m-noise-record/asr_voice.xlsx"

def returndictinfobyexcel(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    allname = workbook.get_sheet_names()
    dictinfo={}
    for one_sheet in allname:
        table = workbook.get_sheet_by_name(one_sheet)
        # 获取总行数
        nrows = table.max_row + 1
        for i in range(1, nrows):
            wav_value = str(table.cell(i,1).value) # wav文件值
            if wav_value=="None":
                continue
            print(wav_value)
            result_value = table.cell(i, 2).value  # 预期结果值
            split_value = wav_value.split("_")
            print(split_value)
            middle_value = str(split_value[2]) + "_" + str(split_value[3])
            middle_value = middle_value.replace(".wav", "")
            if middle_value not in dictinfo:
               dictinfo[middle_value]=result_value

    return  dictinfo


def read_txt():
    dict_info=returndictinfobyexcel(excel_path)
    ofobj=OperationFile()
    alllines=ofobj.read_fileinfo(txt_path)
    print(dict_info)
    new_lines=""
    for i in alllines:
        i=i.replace("\n","")
        replace_value=i.replace("3m-45degree-quiet/","")
        print(replace_value)
        split_value = replace_value.split("_")
        print(split_value)
        middle_value = str(split_value[2]) + "_" + str(split_value[3])
        middle_value = middle_value.replace(".wav", "")

        if dict_info[middle_value] != None:
            except_info = str(dict_info[middle_value]).strip()
            print(except_info)
            new_lines=new_lines+replace_value+ "    "+ except_info+"\n"

        print(source_dir+i+".pcm")
        print(copy_dir+i.replace("3m-45degree-quiet/","")+".pcm")
        ofobj.start_copy_file(source_dir+i+".pcm",copy_dir+i.replace("3m-45degree-quiet/","")+".pcm")

    ofobj.write_fileinfo(new_txt_path,new_lines)


if __name__=="__main__":
    read_txt()

