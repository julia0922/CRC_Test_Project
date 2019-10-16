#!/usr/bin/env python
# coding=utf-8
### websocket模块https://pypi.python.org/pypi/websocket-client
### wavfile是一个speex压缩的ogg文件,文件格式是Ogg data, Speex audio
import websocket
from gevent import os
from websocket import ABNF
import time
import _thread
import json
import xlrd
from xlutils.copy import copy
import  openpyxl
import os
import pandas as pd

from multiprocessing import Process

# sit环境执行地址
product_id = '278575321' #填入产品Id
url = 'wss://dds-hb.dui.ai/dds/v1/test?serviceType=websocket&productId=' + product_id
# 进程数
threadNum=10
# 是否去除第一行
isRemove_firstline=True
# 用例文件xlsx
excel_path="E:/audiofile/100人粤语数据-第二批交付（普通话2人）/test.xlsx"
# 音频文件所在的位置
cmd_path="E:/audiofile/100人粤语数据-第二批交付（普通话2人）/data/wav/"
# 多进程执行完成后用例保存的位置、识别率汇总文件保存的位置
new_root_path="D:/100人粤语-采集-批次临时普通话-执行结果1/"
#日志文件目录
log_rootdir="D:/logroot/"

# 打开excel所在的文件路径，获取excel内容
data=xlrd.open_workbook(excel_path)
table=None
def on_message(ws, message):
   print ('Message>>>>>>>>>>>>>>>>>>>>>>>>')
   print(message)
   mJson = json.loads(message)
   text = ""
   global currentindex,new_sheet
   print(currentindex)
   index = currentindex
   if 'eof' in message:
       text = mJson['text']
       text=text.replace(" ","")
       print(text.replace(" ",""))
       new_sheet.cell(index,3).value=text
       except_value=str(new_sheet.cell(index,2).value).strip()
       print(except_value)
       print(except_value.lower())
       if (except_value.lower() == text.lower()):
           new_sheet.cell(index,4).value="True"
       else:
           new_sheet.cell(index, 4).value = "False"

def on_error(ws, error):
    print ('error>>>>>>>>>>>>>>>>>>>>>>')
    print (error)
    writelog(log_wholedpath, error)
    print ('error<<<<<<<<<<<<<<<<<<<<<<')

def on_close(ws):
    print("文件路径:",filepath)
    print("isEnd" ,isEnd)
    if isEnd==False:
        writelog(log_wholedpath,"未执行完成.......")
        #func_one(filepath)
    print ("### closed ###")

def on_open(ws):
    def run(*args):
        global currentindex, filepath, new_sheet,log_wholedpath,isEnd
        basename = os.path.basename(filepath)
        log_wholedpath = log_rootdir + str(basename) + ".log"
        isEnd=False
        mywb = openpyxl.load_workbook(filepath)
        list = filepath.split("/")
        print(list[len(list) - 2])
        s_name=list[len(list) - 2]
        print("**********",s_name)
        new_sheet = mywb.get_active_sheet()
        try:
            for i in range(2, new_sheet.max_row+1):
                currentindex=i
                wav_value = new_sheet.cell(i, 1).value  # wav文件值
                print("地址文件名称....",wav_value)
                if wav_value=="" or wav_value==None:
                    continue

                if wav_value.find(".wav")==-1:
                    wav_value=wav_value+".wav"

                wav_path=cmd_path+"/"+s_name+"/"+wav_value
                print(wav_path)
                content = {
                    "topic": "recorder.stream.start",
                    "recordId": "21354578ijhgbvdrt01",
                    "audio": {
                        "audioType": "wav",
                        "sampleRate": 16000,
                        "channel": 1,
                        "sampleBytes": 2
                    },
                    "asrParams": {
                        "enableVAD": False,
                        "realBack": False,
                        "toneEnable": True
                    },
                    "aiType": "asr"
                }
                ws.send(json.dumps(content))
                step = 3200  # 如果audioType是wav，此处需要修改为3200
                with open(wav_path, 'rb') as f:
                    while True:
                        wave_data = f.read(step)
                        if wave_data:
                            ws.send(wave_data, ABNF.OPCODE_BINARY)
                        if len(wave_data) < step:
                            break
                        time.sleep(0.1)
                ws.send('', ABNF.OPCODE_BINARY)
                time.sleep(0.6)
            isEnd = True
            mywb.save(filepath)
            mywb.close()
        except Exception as e:
            writelog(log_wholedpath,"发生异常了.....")
        finally:
            #mywb.save(filepath)
            #mywb.close()
            ws.close()
    _thread.start_new_thread(run, ())


def writelog(logpath,message):
    with open(logpath, mode='w') as write:
        write.write(message)


def func_one(var_filepath):
    global filepath
    filepath=var_filepath


    websocket.enableTrace(True)
    ws = websocket.WebSocketApp(url,
                                on_message=on_message,
                                on_error=on_error,
                                on_close=on_close)
    ws.on_open = on_open
    ws.run_forever()

def getargvalue(index,step_list):
    fistagr=0
    secondarg=0
    for i in range(0, index):
        fistagr=fistagr+step_list[i]

    for i in range(0, index+1):
        secondarg=secondarg+step_list[i]
    return fistagr,secondarg


def getavage(total_sheet):
    total=total_sheet
    j = total / threadNum  # 除数,python除法不会四舍五入，52
    q = total % threadNum
    print("j",j)
    print("q",q)

    # k = m % n  # 余数要再被依次分配给n，直到分完,8
    step_list = []
    for i in range(threadNum):
        step_list.append(int(j))

    if q > 0:  # 余数>0时, 除数再依次分配余数，直到分完
        for i in range(q):
            step_list[i] += 1

    return step_list


def count_recognitionrate():
    totalwb=openpyxl.Workbook()
    mywb = openpyxl.Workbook()
    new_sheet = mywb.create_sheet(index=0, title='识别率')
    row1 = ["音频", "案例总数", "成功数", "失败数", "识别率"]
    new_sheet.append(row1)

    for dirname in os.listdir(new_root_path):
        if os.path.isdir(new_root_path+dirname)==False:
            continue
        file_root_path=new_root_path+dirname+"/output.xlsx"
        data = openpyxl.load_workbook(file_root_path)
        table =data.worksheets[0]
        sum=sucesscount=failcount=0
        sucess_rate_value=0.0
        row_list=[]
        for row in range(2,table.max_row+1):
            result_value=str(table.cell(row,4).value)
            sum=sum+1
            if result_value=="True" or result_value=="1" or result_value=="TRUE":
                sucesscount=sucesscount+1
            else:
                failcount=failcount+1
            if sum != 0:
                sucess_rate = (sucesscount / sum) * 100
                sucess_rate_value = "{:.2f}%".format(sucess_rate)

            one_row=[str(table.cell(row,1).value),str(table.cell(row,2).value),str(table.cell(row,3).value),result_value]
            row_list.append(one_row)
        # 汇总所有的用例文件到一个excel中。
        t_sheet = totalwb.create_sheet(title=dirname)
        t_sheet.append(["Recognition success rate",sucess_rate_value])
        t_sheet.append(["wav","expect","asr","recognition result"])
        for r_i in row_list:
            t_sheet.append(r_i)

        # 统计所有的sheet的识别率情况。
        new_sheet.append([dirname, sum, sucesscount, failcount, sucess_rate_value])
    mywb.save( new_root_path+"total_rate.xlsx")
    totalwb.save(new_root_path+"汇总.xlsx")

# 合并目录下的excel文件到一个excel中
def contact_excel():
    for dirname in os.listdir(new_root_path):
        index = 0
        tempExcel = None
        file_root_path = new_root_path + dirname
        if os.path.isfile(file_root_path):
            continue

        if os.path.exists(file_root_path + "/" + 'output.xlsx'):
            os.remove(file_root_path + "/" + 'output.xlsx')

        for filename in os.listdir(file_root_path):
            filePath = file_root_path+"/" + filename
            print(filePath)
            if index == 0:
                tempExcel = pd.read_excel(filePath)
            else:
                a2 = pd.read_excel(filePath)
                tempExcel = pd.concat([tempExcel, a2])
                print(tempExcel)
            index = index + 1

        writer = pd.ExcelWriter(file_root_path +"/"+ 'output.xlsx')
        tempExcel.to_excel(writer,index=None)
        writer.save()
        writer.close()

def writeExcelTodir(fileinfo,data):
    writer = pd.ExcelWriter(fileinfo)
    print(data)
    data.to_excel(writer,index=None)
    writer.save()
    writer.close()

if __name__ == "__main__":
    if os.path.exists(new_root_path)==False:
        os.mkdir(new_root_path)
    if os.path.exists(log_rootdir) == False:
        os.mkdir(log_rootdir)

    names = data.sheet_names()   # 获取所有的sheet
    print(names)

    for name in names:
        print(name)
        if isRemove_firstline==True:
            table = pd.read_excel(excel_path,sheet_name=name,skiprows=1)
        else:
            table = pd.read_excel(excel_path, sheet_name=name)

        # 获取总行数
        nrows=len(table)
        processlist = []
        if nrows < threadNum:
            steplist = [nrows]
        else:
            steplist = getavage(nrows)

        for index in range(0, len(steplist)):
            if index == 0:
                startindex = 0
                endindex = steplist[index]
            else:
                newsteplist = steplist[:index]
                newsteplist2 = steplist[:index +1]
                startindex = sum(newsteplist)
                endindex = sum(newsteplist2)

            print("startindex:,", startindex)
            print("endindex:,", endindex)
            temp_data=table[startindex:endindex]
            one_sheet_dir=new_root_path+name+"/"
            one_sheet_name=name+"_"+str(startindex)+"_"+str(endindex)+".xlsx"
            if os.path.exists(one_sheet_dir) == False:
                os.mkdir(one_sheet_dir)

            temp_fileinfo=one_sheet_dir+one_sheet_name
            print(temp_fileinfo)
            if os.path.exists(temp_fileinfo)==False:
                writeExcelTodir(temp_fileinfo,temp_data)
            p_one = Process(target=func_one, args=(temp_fileinfo,))
            processlist.append(p_one)
            p_one.start()

        for i in processlist:
            i.join()

    print("执行完成")

    contact_excel()
    # 计算识别率
    count_recognitionrate()









